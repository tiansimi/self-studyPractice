# 一个以.xlsx为扩张名的excel文件打开后叫工作簿workbook
# 正在操作的这张表单被认为是活跃的active sheet。
# 每张表单有行和列，行号1、2、3…，列号A、B、C...。
# 在某一个特定行和特定列的小格子叫单元格cell。
from openpyxl import Workbook
import openpyxl
import datetime

# wb = Workbook()

# 打开Excel文件，获取工作表对象
wb1 = openpyxl.load_workbook('example.xlsx')
ws = wb1.active  # 当前活跃的工作簿

# mySheet = wb1.create_sheet('mySheet')
# mySheet.title = 'newSheet'
sheets = wb1.sheetnames
print(sheets)
cc = ws['A1']  # 获取A1这个单元格对象
c = ws['A4'].value
ws['A2'] = datetime.datetime.now()
print(c)
print(ws['A2'].value)  # 获取单元格内容
sheet1 = wb1.get_sheet_by_name('newSheet')  # 获取指定表单
c1 = sheet1['A4'].value
print("第二个sheet的A4是:%s" % c1)
sheet3 = wb1['newSheet1']  # 获取指定表单
c2 = sheet3['A4'].value
print('第三个sheet的A4是：%s' % c2)
sheet1.append(['‘This is A1', '‘This is B1’', '‘This is C1’'])
sheet1.append(['‘This is A1', '‘This is B1’', '‘This is C1’'])
print(sheet1['A1'].value)
sheet1.cell(row=6, column=6).value = 99999
print(sheet1.rows)
print('最小行为：%s' % sheet1.min_row)
print(sheet1.cell(1, 3).value)
minrow = sheet1.min_row  # 最小行
maxrow = sheet1.max_row  # 最大行
mincol = sheet1.min_column  # 最小列
maxcol = sheet1.max_column  # 最大列
# 按行读取
print('按行读取')
for i in range(minrow, maxrow + 1):
    for j in range(mincol, maxcol + 1):
        print(sheet1.cell(i, j).value, end='')  # 加end='' 是为了打印后不换行
# 按行读取第二种方法（可以限制行和列）
data_all = sheet1.iter_rows(min_row=1, max_row=6, min_col=1, max_col=3)
data_all_tuple = tuple(data_all)
for rows in data_all_tuple:
    for cells in rows:
        print('按行读取第二种方法%s' % cells.value)

# 做个人吧
wb1.save('example.xlsx')
