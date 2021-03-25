from openpyxl import load_workbook


class Doexcel:

    def get_data(self, file_name, sheet_name):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            row_data['url'] = sheet.cell(i, 1).value
            row_data['data'] = sheet.cell(i, 2).value
            row_data['title'] = sheet.cell(i, 3).value
            row_data['http_method'] = sheet.cell(i, 4).value
            test_data.append(row_data)
        return test_data


if __name__ == "__main__":
    test_data = Doexcel().get_data('url表.xlsx', 'owner_info')
